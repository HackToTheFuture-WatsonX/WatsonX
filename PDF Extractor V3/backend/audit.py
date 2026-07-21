"""
audit.py — AuditResource feature for PDF Extractor V3.

Flattens extracted background-check reports into the persisted `audit_records`
table (the real-time source of truth), and serves the Audit page (master list),
server-side Excel export, user overrides, and audit-driven Insights stats.

The persisted table is written at extraction time (extractor.py calls
`flatten_and_store`). A one-time backfill (`backfill_from_json`) populates it
from JSON files already on disk.
"""
import io
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

import db
from config import ai_json_dir

router = APIRouter(prefix="/api/audit", tags=["audit"])

# Ordered AuditResource column labels (the audit_resource view returns these).
AUDIT_COLUMNS = [
    "S/N", "Candidate Name", "Initiation Date", "Final Report Sent Date",
    "Supplementary Report Sent Date", "Overall BGV Result",
    "E1 (most recent)", "E2", "E3", "E4", "E5", "REF 1", "REF 2",
    "Adverse Media Check", "Global Sanctions", "Bankruptcy Check",
    "Financial/Credit Check", "Directorship Check (DTI Only)",
    "Civil Litigation Check", "Professional License Qualification",
    "Social Media Screening", "Name", "Onboarding Date",
    "Background Check Date", "isCompliant",
]

# Map an "other_checks"/database check_name -> audit_records column.
_CHECK_TO_COLUMN = {
    "adverse media check": "adverse_media",
    "global sanctions": "global_sanctions",
    "bankruptcy check": "bankruptcy",
    "financial/credit check": "financial_credit",
    "directorship check": "directorship",
    "directorship check (dti only)": "directorship",
    "civil litigation check": "civil_litigation",
    "professional license qualification": "professional_license",
    "social media screening": "social_media",
}


# ── Flatten mapping ───────────────────────────────────────────────────────────

def flatten_structured(structured: dict) -> dict:
    """Map a build_structured_json() dict → audit_records column dict.

    Employers (E1–E5) and references (REF 1/2) come from the summary lists in
    the order provided (most recent first). Database/other checks are keyed by
    check_name (case-insensitive).
    """
    summary = structured.get("report_summary", {}) or {}

    emp = [ (e.get("employer") or "").strip()
            for e in summary.get("employment_check_summary", []) or [] ]
    emp += [""] * (5 - len(emp))

    refs = [ (r.get("referee") or "").strip()
             for r in summary.get("professional_reference_summary", []) or [] ]
    refs += [""] * (2 - len(refs))

    out = {
        "candidate_name":            (summary.get("subject_name") or "").strip(),
        "initiation_date":           (summary.get("case_received") or "").strip(),
        "final_report_date":         (summary.get("delivery_date") or "").strip(),
        "supplementary_report_date": "",
        "overall_bgv_result":        (summary.get("overall_status") or "").strip(),
        "e1": emp[0], "e2": emp[1], "e3": emp[2], "e4": emp[3], "e5": emp[4],
        "ref1": refs[0], "ref2": refs[1],
        "adverse_media": "", "global_sanctions": "", "bankruptcy": "",
        "financial_credit": "", "directorship": "", "civil_litigation": "",
        "professional_license": "", "social_media": "",
    }

    # Database check results: prefer report_summary.database_check_summary, then
    # fall back to structured.other_checks. Keyed by check_name.
    def _apply(check_name: str, result: str):
        col = _CHECK_TO_COLUMN.get((check_name or "").strip().lower())
        if col and result:
            out[col] = result.strip()

    for entry in summary.get("database_check_summary", []) or []:
        _apply(entry.get("check", ""), entry.get("result") or entry.get("status") or "")
    for oc in structured.get("other_checks", []) or []:
        col = _CHECK_TO_COLUMN.get((oc.get("check_name") or "").strip().lower())
        if col and not out.get(col):
            out[col] = (oc.get("result") or oc.get("status") or "").strip()

    return out


def flatten_and_store(structured: dict, source_json: str = "") -> str:
    """Flatten + upsert into audit_records. Returns the ref_number used."""
    summary = structured.get("report_summary", {}) or {}
    ref = (summary.get("case_reference") or "").strip()
    if not ref:
        return ""
    fields = flatten_structured(structured)
    fields["source_json"] = source_json or ""
    db.audit_record_upsert(ref, fields)
    return ref



# ── Backfill ──────────────────────────────────────────────────────────────────

def backfill_from_json() -> dict:
    """Populate audit_records from every JSON file on disk. Returns a summary.

    Newest files are processed last so that, on a duplicate ref_number, the most
    recently extracted report wins (INSERT OR REPLACE). User overrides are never
    touched. Returns {scanned, stored, skipped, errors}.
    """
    json_dir = ai_json_dir()
    scanned = stored = skipped = errors = 0
    if not json_dir.exists():
        return {"scanned": 0, "stored": 0, "skipped": 0, "errors": 0}

    files = sorted(json_dir.rglob("*.json"), key=lambda p: p.stat().st_mtime)
    for jp in files:
        scanned += 1
        try:
            structured = json.loads(jp.read_text(encoding="utf-8"))
            ref = flatten_and_store(structured, source_json=str(jp))
            if ref:
                stored += 1
            else:
                skipped += 1
        except Exception:
            errors += 1
    return {"scanned": scanned, "stored": stored, "skipped": skipped, "errors": errors}


# ── Stats ─────────────────────────────────────────────────────────────────────

def _period_key(raw: str, period: str) -> str:
    """Bucket an ISO-ish date string into a day/week/month/year label."""
    raw = (raw or "").strip()
    if not raw:
        return "(unknown)"
    dt = None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y", "%Y/%m/%d"):
        try:
            dt = datetime.strptime(raw[:len(fmt) + 4], fmt)
            break
        except (ValueError, TypeError):
            continue
    if dt is None:
        try:
            dt = datetime.fromisoformat(raw)
        except (ValueError, TypeError):
            return "(unknown)"
    if period == "day":
        return dt.strftime("%Y-%m-%d")
    if period == "week":
        return f"Week {dt.isocalendar()[1]:02d}, {dt.year}"
    if period == "year":
        return str(dt.year)
    return dt.strftime("%b %Y")


def compute_stats(period: str = "month") -> dict:
    """Compute audit-driven Insights stats from the audit_resource view.

    Returns:
      - stats: total, compliant, non_compliant, compliant_with_onboarding
      - onboarding_chart: onboarding-date counts bucketed by the given period
    """
    rows = db.audit_resource_all()
    total = len(rows)
    compliant = sum(1 for r in rows if str(r.get("isCompliant")).lower() == "true")
    non_compliant = total - compliant
    compliant_with_onboarding = sum(
        1 for r in rows
        if str(r.get("isCompliant")).lower() == "true"
        and (r.get("Onboarding Date") or "").strip()
    )

    buckets: dict[str, int] = defaultdict(int)
    for r in rows:
        onboard = (r.get("Onboarding Date") or "").strip()
        if onboard:
            buckets[_period_key(onboard, period)] += 1
    onboarding_chart = [
        {"period": k, "count": v} for k, v in sorted(buckets.items())
    ]

    return {
        "stats": {
            "total": total,
            "compliant": compliant,
            "non_compliant": non_compliant,
            "compliant_with_onboarding": compliant_with_onboarding,
        },
        "onboarding_chart": onboarding_chart,
        "period": period,
    }


# ── Excel export ──────────────────────────────────────────────────────────────

def _export_excel() -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = db.audit_resource_all()
    wb = Workbook()
    ws = wb.active
    ws.title = "AuditResource"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    ws.append(AUDIT_COLUMNS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append([r.get(col, "") for col in AUDIT_COLUMNS])

    for idx, col in enumerate(AUDIT_COLUMNS, start=1):
        width = min(max(len(col) + 2, 12), 40)
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else
                             "A" + chr(64 + idx - 26)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"AuditResource_{ts}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )



# ── REST endpoints ──────────────────────────────────────────────────────────

class OverrideRequest(BaseModel):
    ref_number: str
    candidate_name: str | None = None
    onboarding_date: str | None = None
    background_check_date: str | None = None
    is_compliant: str | None = None


@router.get("")
def audit_list():
    """Return the full AuditResource master list (labelled columns)."""
    return {"columns": AUDIT_COLUMNS, "rows": db.audit_resource_all()}


@router.post("/override")
def audit_override(req: OverrideRequest):
    fields = {
        "candidate_name": req.candidate_name,
        "onboarding_date": req.onboarding_date,
        "background_check_date": req.background_check_date,
        "is_compliant": req.is_compliant,
    }
    db.audit_override_upsert(req.ref_number, fields)
    return {"status": "ok"}


@router.post("/backfill")
def audit_backfill():
    return backfill_from_json()


@router.get("/stats")
def audit_stats(period: str = "month"):
    return compute_stats(period)


@router.get("/export")
def audit_export():
    """Server-side Excel export of the full AuditResource columns."""
    return _export_excel()  # filled below
